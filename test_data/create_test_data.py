import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# 共通設定
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.border = THIN_BORDER
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# --- テスト用の受講生データ（30名、共通で使う） ---
MEMBERS = [
    ("山田太郎", "2025-01-15", "コーチング", "active"),
    ("佐藤花子", "2025-02-01", "セラピスト", "active"),
    ("鈴木一郎", "2025-01-20", "コンサルタント", "active"),
    ("田中美咲", "2025-03-10", "ヨガインストラクター", "active"),
    ("高橋健太", "2025-02-15", "パーソナルトレーナー", "active"),
    ("伊藤めぐみ", "2025-04-01", "カウンセラー", "active"),
    ("渡辺大輔", "2025-01-10", "経営コンサル", "active"),
    ("中村さくら", "2025-05-01", "ネイリスト", "active"),
    ("小林陽子", "2025-03-20", "料理教室講師", "active"),
    ("加藤翔太", "2025-06-01", "Webデザイナー", "active"),
    ("吉田真理", "2025-02-20", "英会話講師", "active"),
    ("松本大地", "2025-04-15", "整体師", "active"),
    ("木村あかり", "2025-01-05", "フォトグラファー", "active"),
    ("林直樹", "2025-07-01", "ファイナンシャルプランナー", "active"),
    ("清水由美", "2025-03-01", "エステティシャン", "active"),
    ("森田拓也", "2025-05-15", "プログラマー", "active"),
    ("藤本香織", "2025-08-01", "フラワーアレンジメント講師", "active"),
    ("岡田智子", "2025-02-10", "占い師", "active"),
    ("長谷川悟", "2025-06-15", "動画クリエイター", "active"),
    ("石井美穂", "2025-04-20", "ピラティスインストラクター", "active"),
    ("前田康介", "2025-09-01", "コーチング", "active"),
    ("小川理恵", "2025-07-15", "アロマセラピスト", "active"),
    ("村上浩二", "2025-01-25", "税理士", "retired"),
    ("近藤裕子", "2025-03-15", "ハンドメイド作家", "retired"),
    ("坂本光", "2025-02-05", "パーソナルスタイリスト", "retired"),
    ("遠藤真一", "2025-05-20", "コピーライター", "retired"),
    ("青木麻衣", "2025-08-10", "ダンスインストラクター", "active"),
    ("西村健", "2025-06-20", "マーケター", "active"),
    ("福田さやか", "2025-09-15", "管理栄養士", "active"),
    ("太田圭介", "2025-10-01", "スポーツトレーナー", "active"),
]

ACTIVE_MEMBERS = [(n, d, j, s) for n, d, j, s in MEMBERS if s == "active"]
RETIRED_MEMBERS = [(n, d, j, s) for n, d, j, s in MEMBERS if s == "retired"]

# --- 1. 受講生の名簿 ---
def create_meibo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受講生名簿"
    ws.append(["名前", "入会日", "職種", "ステータス", "メールアドレス", "LINE名"])
    for name, date, job, status in MEMBERS:
        email = f"{name}@example.com"
        line_name = name
        status_label = "在籍中" if status == "active" else "退会済"
        ws.append([name, date, job, status_label, email, line_name])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "受講生の名簿.xlsx"))

# --- 2. 入会前の事前課題の回答データ ---
def create_jizen_kadai():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "事前課題"
    ws.append(["回答日", "名前", "現在の職種", "経験年数", "月商（万円）", "目標月商（万円）",
               "一番の悩み", "過去に試したこと", "入会のきっかけ", "週に使える時間（時間）"])

    mondai = ["集客ができない", "成約率が低い", "リピートが取れない", "単価を上げられない",
              "時間管理ができない", "発信が続かない", "差別化ができない"]
    tameshita = ["SNS発信", "ブログ", "チラシ配布", "紹介営業", "広告出稿", "セミナー開催", "特になし"]
    kikkake = ["紹介", "SNSで見た", "セミナーに参加して", "ブログを読んで", "YouTube"]

    for name, date, job, status in MEMBERS:
        ws.append([
            date, name, job,
            random.randint(1, 15),
            random.choice([0, 5, 10, 15, 20, 30, 50]),
            random.choice([30, 50, 100, 150, 200, 300]),
            random.choice(mondai),
            random.choice(tameshita),
            random.choice(kikkake),
            random.choice([3, 5, 7, 10, 15, 20])
        ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "入会前の事前課題の回答データ.xlsx"))

# --- 3. 既存コンテンツの一覧 ---
def create_contents():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "コンテンツ一覧"
    ws.append(["コンテンツID", "タイトル", "カテゴリ", "形式", "時間（分）", "URL", "公開日"])

    contents = [
        ("C001", "SNS集客の基礎", "集客", "動画", 32, "2025-01-10"),
        ("C002", "ターゲット設定ワーク", "集客", "ワークシート", 0, "2025-01-15"),
        ("C003", "LINE公式アカウントの作り方", "集客", "動画", 45, "2025-02-01"),
        ("C004", "Instagram発信テンプレート", "集客", "テンプレート", 0, "2025-02-10"),
        ("C005", "ストーリーズ活用術", "集客", "動画", 28, "2025-02-20"),
        ("C006", "LP（ランディングページ）の作り方", "集客", "動画", 55, "2025-03-01"),
        ("C007", "メルマガ/LINE配信の書き方", "集客", "動画", 38, "2025-03-15"),
        ("C008", "広告運用入門", "集客", "動画", 42, "2025-04-01"),
        ("C009", "SEOブログの書き方", "集客", "動画", 35, "2025-04-15"),
        ("C010", "紹介を生む仕組み作り", "集客", "動画", 30, "2025-05-01"),
        ("C011", "集客導線の全体設計", "集客", "動画", 60, "2025-06-01"),
        ("C012", "無料オファーの作り方", "集客", "ワークシート", 0, "2025-06-15"),
        ("C013", "セールストーク基礎", "成約", "動画", 40, "2025-02-15"),
        ("C014", "体験セッションの進め方", "成約", "動画", 35, "2025-03-10"),
        ("C015", "クロージングの型", "成約", "動画", 30, "2025-05-15"),
        ("C016", "リピート率を上げる顧客フォロー", "リピート", "動画", 25, "2025-04-20"),
        ("C017", "マインドセット：続ける力", "マインド", "動画", 20, "2025-01-20"),
        ("C018", "時間管理術", "マインド", "動画", 22, "2025-03-20"),
        ("C019", "ビジョン設定ワーク", "マインド", "ワークシート", 0, "2025-01-25"),
        ("C020", "差別化コンセプトの作り方", "ブランディング", "動画", 48, "2025-07-01"),
    ]

    for cid, title, cat, fmt, mins, pub_date in contents:
        url = f"https://example.com/contents/{cid.lower()}"
        ws.append([cid, title, cat, fmt, mins if mins > 0 else "", url, pub_date])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "既存コンテンツの一覧.xlsx"))

# --- 4. セミナー満足度アンケート ---
def create_seminar_enquete():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "セミナー満足度"
    ws.append(["回答日", "セミナー日", "セミナーテーマ", "名前", "満足度（1-5）",
               "理解度（1-5）", "実践したいこと", "自由記述"])

    themes = ["集客の基礎", "セールスの型", "LINE活用", "ブランディング",
              "マインドセット", "時間管理", "リピート戦略", "広告入門",
              "コンテンツ作成", "紹介の仕組み", "差別化戦略", "目標設定"]
    jissen = ["SNS発信を毎日する", "LINE登録導線を作る", "体験セッションを改善する",
              "ターゲットを絞り直す", "時間をブロックする", "紹介カードを作る"]
    months = ["2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06",
              "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]

    for i, month in enumerate(months):
        sem_date = f"{month}-15"
        theme = themes[i]
        attendees = random.sample(ACTIVE_MEMBERS, random.randint(15, 22))
        for name, _, _, _ in attendees:
            ws.append([
                sem_date, sem_date, theme, name,
                random.choices([3, 4, 5], weights=[10, 40, 50])[0],
                random.choices([2, 3, 4, 5], weights=[5, 15, 40, 40])[0],
                random.choice(jissen),
                random.choice(["", "", "", "とても参考になりました", "もっと深く知りたい", "具体例が欲しい"])
            ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "セミナー満足度アンケートの回答データ.xlsx"))

# --- 5. グルコン参加後アンケート ---
def create_groupcon_enquete():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "グルコンアンケート"
    ws.append(["回答日", "グルコン日", "名前", "満足度（1-5）",
               "相談した内容", "得られた気づき", "次にやること"])

    soudan = ["集客がうまくいかない", "成約率を上げたい", "発信のネタがない",
              "LINEの使い方がわからない", "単価設定に悩んでいる", "時間が足りない",
              "差別化ができない", "リピートが取れない", "モチベーションが下がっている",
              "何から手をつけていいかわからない"]
    kizuki = ["ターゲットを絞ることの重要性", "まず行動することの大切さ",
              "他の人も同じ悩みを持っていて安心した", "具体的なステップがわかった",
              "自分の強みに気づけた", "優先順位が明確になった"]
    next_action = ["SNS発信を再開する", "LINE導線を見直す", "体験セッションを3件入れる",
                   "ターゲットシートを書き直す", "コンテンツを3本見る", "時間割を作る"]

    for month_num in range(1, 13):
        for session in range(random.randint(2, 4)):
            gc_date = f"2025-{month_num:02d}-{random.choice([5, 10, 15, 20, 25])}"
            attendees = random.sample(ACTIVE_MEMBERS, random.randint(5, 12))
            for name, _, _, _ in attendees:
                ws.append([
                    gc_date, gc_date, name,
                    random.choices([3, 4, 5], weights=[10, 40, 50])[0],
                    random.choice(soudan),
                    random.choice(kizuki),
                    random.choice(next_action)
                ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "グルコン参加後アンケートの回答データ.xlsx"))

# --- 6. 個別セッション満足度アンケート ---
def create_session_enquete():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "セッション満足度"
    ws.append(["回答日", "セッション日", "名前", "担当者", "満足度（1-5）",
               "相談内容", "解決度（1-5）", "感想"])

    tantou = ["講師A", "講師B", "講師C"]
    soudan = ["集客方法について", "セールスの流れについて", "LINE構築について",
              "コンセプト作りについて", "モチベーションについて", "時間管理について",
              "価格設定について", "SNS発信について", "紹介の仕方について"]
    kanso = ["具体的なアドバイスがもらえた", "方向性が明確になった", "背中を押してもらえた",
             "新しい視点が得られた", "すぐに実践できる内容だった", ""]

    for month_num in range(1, 13):
        for name, _, _, status in ACTIVE_MEMBERS:
            if random.random() < 0.4:
                s_date = f"2025-{month_num:02d}-{random.randint(1, 28)}"
                ws.append([
                    s_date, s_date, name,
                    random.choice(tantou),
                    random.choices([3, 4, 5], weights=[5, 35, 60])[0],
                    random.choice(soudan),
                    random.choices([3, 4, 5], weights=[10, 40, 50])[0],
                    random.choice(kanso)
                ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "個別セッション満足度アンケートの回答データ.xlsx"))

# --- 7. 実績アンケート ---
def create_jisseki():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "実績アンケート"
    ws.append(["回答日", "名前", "入会前の月商（万円）", "現在の月商（万円）",
               "成果が出た施策", "成果が出るまでの期間", "一番役に立ったコンテンツ",
               "一番役に立ったサポート", "後輩へのアドバイス"])

    sesaku = ["SNS集客", "LINE配信", "紹介営業", "セミナー集客", "広告運用", "ブログSEO"]
    kikan = ["1ヶ月", "2ヶ月", "3ヶ月", "6ヶ月", "9ヶ月", "12ヶ月"]
    content_fav = ["SNS集客の基礎", "LINE公式アカウントの作り方", "セールストーク基礎",
                   "集客導線の全体設計", "体験セッションの進め方", "差別化コンセプトの作り方"]
    support_fav = ["個別セッション", "グルコン", "月一セミナー", "グルコン＋個別セッション"]
    advice = ["とにかく行動すること", "完璧を求めないこと", "グルコンに毎回出ること",
              "言われたことを素直にやること", "仲間と励まし合うこと"]

    achievers = random.sample(ACTIVE_MEMBERS, 8)
    for name, date, job, _ in achievers:
        before = random.choice([0, 5, 10, 15, 20])
        after = before + random.choice([20, 30, 50, 80, 100, 150])
        ws.append([
            "2025-12-01", name, before, after,
            random.choice(sesaku),
            random.choice(kikan),
            random.choice(content_fav),
            random.choice(support_fav),
            random.choice(advice)
        ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "実績アンケートの回答データ.xlsx"))

# --- 8. 退会申請リスト ---
def create_taikai():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "退会申請"
    ws.append(["申請日", "名前", "入会日", "退会理由", "詳細", "引き止め面談実施", "最終退会日"])

    riyu = ["成果が感じられない", "時間が取れなくなった", "経済的な理由", "他のスクールに移る", "目標を達成した"]
    shousai = [
        "仕事が忙しくなり学習時間が確保できなくなった",
        "半年やったが集客が増えなかった",
        "家庭の事情で出費を抑えたい",
        "知人に紹介された別のスクールに行く",
        "月商50万を達成できたので卒業",
        "何をすればいいかわからなくなった",
        "コンテンツの内容が自分に合わなかった",
    ]

    for name, join_date, job, _ in RETIRED_MEMBERS:
        join = datetime.strptime(join_date, "%Y-%m-%d")
        retire = join + timedelta(days=random.randint(90, 300))
        ws.append([
            (retire - timedelta(days=14)).strftime("%Y-%m-%d"),
            name, join_date,
            random.choice(riyu),
            random.choice(shousai),
            random.choice(["あり", "なし"]),
            retire.strftime("%Y-%m-%d")
        ])
    style_header(ws)
    auto_width(ws)
    wb.save(os.path.join(OUTPUT_DIR, "退会申請リストの回答データ.xlsx"))


if __name__ == "__main__":
    random.seed(42)
    create_meibo()
    create_jizen_kadai()
    create_contents()
    create_seminar_enquete()
    create_groupcon_enquete()
    create_session_enquete()
    create_jisseki()
    create_taikai()
    print("全8ファイルの作成が完了しました")
